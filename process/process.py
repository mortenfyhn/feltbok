#!/usr/bin/env python3
"""Turn field voice notes into rows in the Artsobservasjoner import template.

Pipeline, per .m4a recording (filename carries time + GPS + accuracy):
  1. transcribe locally with faster-whisper (free, offline, Norwegian)
  2. parse the transcript with Claude -> species / count / activity / notes
  3. resolve the GPS to a locality name: nearest entry in your my_localities.csv
     (so it links to an established locality), else a Kartverket place name
  4. append a row to a copy of the "Fugl" sheet

Nothing is ever dropped: a clip Claude is unsure about still produces a row,
marked "Usikker artsbestemming = X" with the raw transcript in the private
comment, so you can fix it on the website's "Kontrollér og publisér" step.

Requires ANTHROPIC_API_KEY in the environment. See README.md.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import shutil
import sys
from dataclasses import dataclass

import requests
from openpyxl import load_workbook

# --- field capture format (must match the Android app's buildName) -----------

FILENAME_RE = re.compile(
    r"^(?P<date>\d{4}-\d{2}-\d{2})T(?P<h>\d{2})-(?P<m>\d{2})-(?P<s>\d{2})"
    r"_lat(?P<lat>NA|-?\d+\.\d+)_lon(?P<lon>NA|-?\d+\.\d+)_acc(?P<acc>NA|\d+)\.m4a$"
)

# Seeded into Whisper to bias decoding toward bird vocabulary. Not exhaustive —
# Claude does the real species correction afterwards.
BIRD_PROMPT = (
    "Fugleobservasjon i felt. Arter: gulspurv, sivspurv, løvsanger, gransanger, "
    "munk, bokfink, bjørkefink, grønnfink, stillits, dompap, rødstrupe, måltrost, "
    "rødvingetrost, gråtrost, svarttrost, blåmeis, kjøttmeis, granmeis, "
    "spettmeis, trekryper, tornirisk, heipiplerke, trepiplerke, linerle, "
    "låvesvale, taksvale, tårnseiler, fiskemåke, gråmåke, sildemåke, "
    "rødnebbterne, makrellterne, tjeld, vipe, storspove, rødstilk, gluttsnipe, "
    "krikkand, stokkand, toppand, kvinand, gravand, grågås, kortnebbgås, "
    "knoppsvane, sangsvane, fiskeørn, musvåk, tårnfalk, spurvehauk."
)


@dataclass
class Clip:
    path: str
    date: str          # DD.MM.YYYY
    time: str          # HH:MM
    lat: float | None
    lon: float | None
    acc: int | None    # GPS accuracy in metres


def parse_clip(path: str) -> Clip | None:
    name = os.path.basename(path)
    m = FILENAME_RE.match(name)
    if not m:
        print(f"  ! skipping unrecognised filename: {name}", file=sys.stderr)
        return None
    d = m.group("date")
    return Clip(
        path=path,
        date=f"{d[8:10]}.{d[5:7]}.{d[0:4]}",
        time=f"{m.group('h')}:{m.group('m')}",
        lat=None if m.group("lat") == "NA" else float(m.group("lat")),
        lon=None if m.group("lon") == "NA" else float(m.group("lon")),
        acc=None if m.group("acc") == "NA" else int(m.group("acc")),
    )


# --- template introspection --------------------------------------------------

def fugl_headers(wb) -> dict[str, int]:
    """Map column name -> 1-based column index from the Fugl header (row 2)."""
    ws = wb["Fugl"]
    return {
        str(c.value).strip(): c.column
        for c in ws[2]
        if c.value is not None
    }


def bird_activities(wb) -> list[str]:
    """The valid Aktivitet values for birds, read from the template's Fugl column."""
    rows = list(wb["Aktivitet"].iter_rows(values_only=True))
    header = rows[1]  # group-header row: Fugl, Kriterie, Amfibier/reptiler, ...
    col = header.index("Fugl")
    return [r[col] for r in rows[2:] if r[col] is not None]


def accuracy_buckets(wb) -> list[int]:
    """Allowed Nøyaktighet radii in metres, ascending."""
    vals = []
    for row in wb["Nøyaktighet"].iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and v.endswith(" m") and v[:-2].strip().isdigit():
                vals.append(int(v[:-2]))
    return sorted(set(vals))


def nearest_bucket(acc: int | None, buckets: list[int]) -> str:
    """Round GPS accuracy UP to an allowed radius — never claim more precision."""
    target = acc if acc is not None else 100  # unknown GPS -> conservative
    for b in buckets:
        if b >= target:
            return f"{b} m"
    return f"{buckets[-1]} m"


# --- transcription -----------------------------------------------------------

def make_transcriber(model_size: str):
    from faster_whisper import WhisperModel

    # A local model path (e.g. the converted NB-Whisper) that isn't built yet
    # falls back to a stock model so the pipeline still runs.
    if os.sep in model_size and not os.path.isdir(model_size):
        print(f"  ! {model_size} not found — falling back to large-v3 "
              f"(run `just nb-whisper` to build it).", file=sys.stderr)
        model_size = "large-v3"

    model = WhisperModel(model_size, device="cpu", compute_type="int8")

    def transcribe(path: str) -> str:
        # Reuse a saved transcript (written next to archived clips) so a re-run
        # over already-processed days costs nothing.
        side = os.path.splitext(path)[0] + ".txt"
        if os.path.exists(side):
            with open(side, encoding="utf-8") as f:
                return f.read().strip()
        segments, _ = model.transcribe(
            path, language="no", initial_prompt=BIRD_PROMPT, vad_filter=True
        )
        return " ".join(s.text for s in segments).strip()

    return transcribe


# --- parsing with Claude -----------------------------------------------------

# Valid Fugl "Alder" values from the template (Norwegian ringing calendar-year
# codes). Constrained as an enum so a stated age never breaks the import.
BIRD_AGES = [
    "Egg", "Pulli", "Adult",
    "1K", "1K+", "2K", "2K+", "2K-", "3K", "3K+", "3K-", "4K", "4K+", "4K-",
    "5K", "5K+", "5K-", "6K", "6K+", "6K-", "7K", "7K+", "7K-",
]

PARSE_TOOL = {
    "name": "record_observation",
    "description": "Record the structured bird observation parsed from the transcript.",
    "input_schema": {
        "type": "object",
        "properties": {
            "species": {
                "type": "string",
                "description": "Norsk navn på arten (Artsobservasjoner-standard), korrigert fra evt. feilhørt tale. Tom streng hvis ingen art nevnes.",
            },
            "count": {
                "type": ["integer", "null"],
                "description": "Antall individer hvis nevnt, ellers null.",
            },
            "activity": {
                "type": ["string", "null"],
                "description": "Nøyaktig én verdi fra den oppgitte aktivitetslista, eller null hvis ingen passer.",
            },
            "sex": {
                "type": ["string", "null"],
                "enum": ["Hann", "Hunn", "Hunnfarget", "I par", None],
                "description": "Kjønn KUN hvis det er eksplisitt nevnt i notatet, ellers null. "
                               "Velg nøyaktig én: Hann, Hunn, Hunnfarget eller I par.",
            },
            "age": {
                "type": ["string", "null"],
                "enum": [*BIRD_AGES, None],
                "description": "Alder KUN hvis det er eksplisitt nevnt, ellers null — ikke gjett. "
                               "Velg nærmeste gyldige kode: f.eks. voksen→Adult, ungfugl/juvenil→1K, "
                               "dununge/reirunge→Pulli, egg→Egg, ellers ringmerkingskoder som 2K, 3K+ osv.",
            },
            "comment": {
                "type": "string",
                "description": "Kun tilleggsinfo som IKKE allerede er fanget i art, antall, aktivitet, "
                               "kjønn eller alder (f.eks. drakt, biotop, retning). Ikke gjenta disse. "
                               "Tom streng hvis ingenting ekstra.",
            },
            "uncertain": {
                "type": "boolean",
                "description": "true hvis artsbestemmelsen er usikker eller transkripsjonen er uklar.",
            },
            "correction": {
                "type": "boolean",
                "description": "true hvis notatet retter en TIDLIGERE observasjon — typisk når det "
                               "starter med «korreksjon» (f.eks. «korreksjon, det var fem grønnfink»). "
                               "I et korreksjonsnotat: fyll KUN ut feltene som faktisk nevnes (resten "
                               "null/tom), men ta med arten hvis den nevnes så riktig observasjon kan finnes.",
            },
        },
        "required": ["species", "count", "activity", "sex", "age", "comment", "uncertain", "correction"],
    },
}


def make_parser(model: str, activities: list[str], species: list[str]):
    import anthropic

    client = anthropic.Anthropic()
    text = (
        "Du tolker norske feltnotater om fugleobservasjoner, transkribert fra tale. "
        "Transkripsjonen kan inneholde feilhørte, uvanlige fuglenavn — korriger til "
        "korrekt norsk artsnavn slik det brukes i Artsobservasjoner. Hvis du er usikker "
        "på arten, eller transkripsjonen er uklar, sett uncertain=true (men gjett likevel "
        "på mest sannsynlige art). Sett sex kun hvis kjønnet er eksplisitt nevnt (Hann, "
        "Hunn, Hunnfarget eller I par), ellers null — ikke gjett. Sett age kun hvis alderen er "
        "eksplisitt nevnt (velg nærmeste gyldige kode, f.eks. voksen→Adult, ungfugl→1K, dununge→Pulli), "
        "ellers null. Hvis notatet starter med «korreksjon» (eller tydelig retter en tidligere "
        "observasjon), sett correction=true og fyll kun ut feltene som faktisk nevnes — ta likevel "
        "med arten hvis den nevnes. Kommentaren skal kun inneholde tilleggsinfo som ikke "
        "allerede er fanget i art, antall eller aktivitet — ikke gjenta disse, og la den "
        "stå tom hvis det ikke er noe ekstra. Velg activity som nøyaktig én verdi fra denne "
        "lista, ellers null:\n"
        + "\n".join(f"- {a}" for a in activities)
    )
    if species:
        text += (
            "\n\nGyldige norske fuglenavn (bruk skrivemåten herfra; velg det nærmeste "
            "hvis transkripsjonen er feilhørt):\n"
            + ", ".join(species)
        )
    system = [{"type": "text", "text": text, "cache_control": {"type": "ephemeral"}}]

    def parse(transcript: str) -> dict:
        if not transcript.strip():
            return {"species": "", "count": None, "activity": None,
                    "sex": None, "age": None, "comment": "", "uncertain": True,
                    "correction": False}
        resp = client.messages.create(
            model=model,
            max_tokens=512,
            system=system,
            tools=[PARSE_TOOL],
            tool_choice={"type": "tool", "name": "record_observation"},
            messages=[{"role": "user", "content": f"Transkripsjon: «{transcript}»"}],
        )
        for block in resp.content:
            if block.type == "tool_use":
                return block.input
        return {"species": "", "count": None, "activity": None,
                "sex": None, "age": None, "comment": "", "uncertain": True,
                "correction": False}

    return parse


# --- corrections -------------------------------------------------------------

# A "korreksjon …" note revises an earlier observation rather than adding a new
# one (e.g. an updated count). It targets the most recent prior observation of
# the same species — or, if it names no species, simply the previous one — and
# overwrites only the fields it actually restates.

def match_correction(prior: list[dict], corr: dict):
    """The earlier obs a correction revises, or None. `prior` is the list of
    already-seen, not-yet-consumed observations in chronological order."""
    candidates = [o for o in prior if not o.get("consumed") and (o["obs"].get("species") or "")]
    sp = (corr.get("species") or "").strip().lower()
    if sp:
        for o in reversed(candidates):
            if (o["obs"].get("species") or "").lower() == sp:
                return o
    return candidates[-1] if candidates else None


def merge_correction(target: dict, corr: dict) -> dict:
    """Overwrite only the fields the correction restates (Replace fields stated)."""
    merged = dict(target)
    for f in ("count", "activity", "sex", "age"):
        if corr.get(f) is not None:
            merged[f] = corr[f]
    for f in ("species", "comment"):
        if (corr.get(f) or "").strip():
            merged[f] = corr[f]
    return merged


# --- locality resolution -----------------------------------------------------

# The import links a row to an existing Artsobservasjoner locality when the name
# matches one *near* the coordinates (verified empirically: coordinates even
# disambiguate same-named localities nationwide). So for your regular spots we
# emit their established name from my_localities.csv; the coordinates then pin
# the right one. Somewhere new, we fall back to a Kartverket place name, which
# creates a fresh point locality the script flags for you to review/rename.


def load_localities(path: str) -> list[tuple[str, float, float]]:
    """Read my_localities.csv (name,lat,lon). Blank/`#` lines ignored."""
    if not path or not os.path.exists(path):
        return []
    out = []
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.reader(f):
            if not r or r[0].lstrip().startswith("#"):
                continue
            if r[0].strip().lower() in ("name", "navn"):  # header
                continue
            try:
                out.append((r[0].strip(), float(r[1]), float(r[2])))
            except (IndexError, ValueError):
                print(f"  ! skipping bad localities row: {r}", file=sys.stderr)
    return out


def load_species(path: str) -> list[str]:
    """Read bird_species.csv (first column = Norwegian name). `#`/blank skipped."""
    if not path or not os.path.exists(path):
        return []
    out = []
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.reader(f):
            if not r or r[0].lstrip().startswith("#"):
                continue
            name = r[0].strip()
            if name and name.lower() not in ("name", "navn", "art"):
                out.append(name)
    return out


def _haversine_m(lat1, lon1, lat2, lon2) -> float:
    r = 6_371_000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def nearest_locality(lat, lon, localities, radius_m):
    """(name, lat, lon) of the closest known locality within radius_m, or None.

    The locality's OWN coordinate is returned (not the observation's GPS) so that
    every observation at the same locality lands on the exact same point and the
    import groups them under one locality instead of minting one per fix.
    """
    best, best_d = None, radius_m
    for name, nlat, nlon in localities:
        d = _haversine_m(lat, lon, nlat, nlon)
        if d <= best_d:
            best, best_d = (name, nlat, nlon), d
    return best


# --- reverse geocoding -------------------------------------------------------

# Coarse administrative names make poor locality labels; prefer something local.
_GEO_SKIP = {"Fylke", "Kommune", "Annen administrativ inndeling"}


def reverse_geocode(lat: float, lon: float):
    """(name, lat, lon) of the nearest Kartverket place — its representation
    point, not the observation's GPS, so repeated visits to the same place share
    one coordinate and group under one locality."""
    try:
        r = requests.get(
            "https://api.kartverket.no/stedsnavn/v1/punkt",
            params={"nord": lat, "ost": lon, "koordsys": 4258,
                    "radius": 2000, "treffPerSide": 20, "utkoordsys": 4258},
            timeout=15,
        )
        r.raise_for_status()
        names = r.json().get("navn", [])
        names.sort(key=lambda n: n.get("meterFraPunkt", 1e9))
        for n in names:
            if n.get("navneobjekttype") in _GEO_SKIP:
                continue
            sn = n.get("stedsnavn") or []
            if sn:
                rp = n.get("representasjonspunkt") or {}
                return sn[0]["skrivemåte"], rp.get("nord", lat), rp.get("øst", lon)
    except Exception as e:
        print(f"  ! geocode failed ({e}); using coordinates", file=sys.stderr)
    return f"{lat:.5f}, {lon:.5f}", lat, lon


# --- output ------------------------------------------------------------------

def pick(cols: dict[str, int], *names: str) -> str:
    """First of `names` present in the sheet (handles v2.20 vs v3.0 renames)."""
    for n in names:
        if n in cols:
            return n
    return names[0]


def write_row(ws, row: int, cols: dict[str, int], values: dict[str, object]) -> None:
    for name, value in values.items():
        if value is None or value == "":
            continue
        if name not in cols:
            continue
        ws.cell(row=row, column=cols[name], value=value)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("recordings", help="folder of .m4a voice notes from the app")
    ap.add_argument("-t", "--template", default="docs/artsobs-template-v3.0.xlsx",
                    help="Artsobservasjoner import template (default: %(default)s)")
    ap.add_argument("-o", "--output", default="observasjoner.xlsx",
                    help="output workbook (default: %(default)s)")
    ap.add_argument("--localities", default="my_localities.csv,localities_gazetteer.csv",
                    help="comma-separated CSVs of localities: name,lat,lon "
                         "(default: %(default)s — your list plus the harvested gazetteer)")
    ap.add_argument("--locality-radius", type=float, default=200.0,
                    help="metres within which to match a known locality (default: %(default)s)")
    ap.add_argument("--species", default="bird_species.csv",
                    help="Norwegian bird checklist to validate against (default: %(default)s)")
    ap.add_argument("--whisper-model", default="models/nb-whisper-large-ct2",
                    help="faster-whisper model: a size (e.g. large-v3) or a local "
                         "path like the NB-Whisper build (default: %(default)s)")
    ap.add_argument("--claude-model", default="claude-sonnet-4-6",
                    help="Claude model for parsing (default: %(default)s). "
                         "claude-opus-4-8 corrects rarer species better at higher cost.")
    ap.add_argument("--keep", action="store_true",
                    help="leave recordings in place instead of moving them to "
                         "processed/ (and skipped/) after a successful run")
    args = ap.parse_args()

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("error: set ANTHROPIC_API_KEY", file=sys.stderr)
        return 1

    clips = [c for c in (
        parse_clip(os.path.join(args.recordings, f))
        for f in sorted(os.listdir(args.recordings)) if f.endswith(".m4a")
    ) if c]
    if not clips:
        print("No recognisable .m4a recordings found.", file=sys.stderr)
        return 1
    print(f"Found {len(clips)} recording(s).")

    wb = load_workbook(args.template)
    cols = fugl_headers(wb)
    activities = bird_activities(wb)
    buckets = accuracy_buckets(wb)
    ws = wb["Fugl"]

    localities = []
    for path in args.localities.split(","):
        localities += load_localities(path.strip())
    print(f"Loaded {len(localities)} known localit(y/ies) from {args.localities}.")

    species = load_species(args.species)
    species_set = {s.lower() for s in species}
    if species:
        print(f"Loaded {len(species)} bird names from {args.species}.")

    transcribe = make_transcriber(args.whisper_model)
    parse = make_parser(args.claude_model, activities, species)

    # Resolved column names (template renamed these between v2.20 and v3.0).
    col_public = pick(cols, "Merknad (synlig for alle)", "Kommentar (synlig for alle)")
    col_private = pick(cols, "Privat merknad (kun synlig for deg selv)",
                       "Privat kommentar (kun synlig for deg selv)")

    out_row = review_row = 3  # row 1 = title, row 2 = headers
    flagged = new_locality = corrections = 0
    ws_review = None  # second workbook, created lazily for unidentified rows
    processed: list[tuple[str, str]] = []  # (path, transcript)
    skipped: list[tuple[str, str]] = []

    # Pass 1: transcribe + parse. A "korreksjon" note folds into the earlier
    # observation it revises instead of becoming its own row, so we collect
    # records first and only write rows once all corrections are applied.
    records: list[dict] = []
    for i, clip in enumerate(clips, 1):
        name = os.path.basename(clip.path)
        print(f"[{i}/{len(clips)}] {name}")
        transcript = transcribe(clip.path)
        print(f"    «{transcript}»")

        # An accidental tap or a note you abandoned has no speech — skip it
        # rather than emitting an empty row.
        if not transcript.strip():
            print("    (tomt – hopper over)")
            skipped.append((clip.path, transcript))
            continue

        obs = parse(transcript)
        processed.append((clip.path, transcript))

        if obs.get("correction"):
            target = match_correction(records, obs)
            if target is not None:
                target["obs"] = merge_correction(target["obs"], obs)
                target["corrections"].append((name, transcript))
                corrections += 1
                print(f"    ↳ korreksjon av {os.path.basename(target['clip'].path)}")
                continue
            print("    ! korreksjon uten match – beholder som egen rad")

        records.append({"clip": clip, "name": name, "transcript": transcript,
                        "obs": obs, "corrections": []})

    # Pass 2: resolve locality and write a row per surviving observation.
    for rec in records:
        clip, name, transcript, obs = rec["clip"], rec["name"], rec["transcript"], rec["obs"]
        sp = obs.get("species") or ""
        off_list = bool(species_set) and sp and sp.lower() not in species_set
        no_gps = clip.lat is None or clip.lon is None
        uncertain = obs.get("uncertain") or no_gps or off_list

        # Locality: a known spot links to the established locality; otherwise a
        # Kartverket name creates a new point that's worth a glance. Either way
        # we output the LOCALITY's coordinate (not the raw GPS) so every
        # observation at the same place groups under one locality.
        is_new_locality = False
        if no_gps:
            locality, lat_out, lon_out = "UKJENT – mangler GPS", None, None
        else:
            match = nearest_locality(clip.lat, clip.lon, localities, args.locality_radius)
            if match is None:
                locality, lat_out, lon_out = reverse_geocode(clip.lat, clip.lon)
                is_new_locality = True
                new_locality += 1
            else:
                locality, lat_out, lon_out = match
            lat_out = round(lat_out, 6)
            lon_out = round(lon_out, 6)

        private = f"[{name}] «{transcript}»"
        if rec["corrections"]:
            tags = "; ".join(f"[{n}] «{t}»" for n, t in rec["corrections"])
            private += f" KORRIGERT: {tags}"
        if no_gps:
            private = "GPS MANGLER. " + private
        elif is_new_locality:
            private = "NY lokalitet (sjekk/omdøp). " + private
        if off_list:
            private = "ART UTENFOR SJEKKLISTE (sjekk). " + private

        row = {
            "Lokalitetsnavn": locality,
            "Nord": lat_out,
            "Øst": lon_out,
            "Nøyaktighet": nearest_bucket(clip.acc, buckets),
            "Fra dato": clip.date,
            "Til dato": clip.date,
            "Fra klokkeslett": clip.time,
            "Til klokkeslett": clip.time,
            "Antall": obs.get("count"),
            "Aktivitet": obs.get("activity"),
            "Kjønn": obs.get("sex"),
            "Alder": obs.get("age"),
            col_public: obs.get("comment"),
        }
        if sp:
            if uncertain:
                flagged += 1
            write_row(ws, out_row, cols, {**row, "Artsnavn": sp, col_private: private,
                                          "Usikker artsbestemming": "X" if uncertain else None})
            out_row += 1
        else:
            # No species at all — keep Artsnavn blank (a "?" breaks the whole
            # import) and route to a separate sheet to fill in by hand.
            if ws_review is None:
                ws_review = load_workbook(args.template)["Fugl"]
            write_row(ws_review, review_row, cols,
                      {**row, col_private: "UIDENTIFISERT – fyll inn art. " + private,
                       "Usikker artsbestemming": "X"})
            review_row += 1

    wb.save(args.output)
    print(f"\nWrote {out_row - 3} importable row(s) to {args.output} "
          f"({flagged} flagged species, {new_locality} new localit(y/ies), "
          f"{corrections} correction(s) folded in; {len(skipped)} empty skipped).")
    if ws_review is not None:
        review_path = re.sub(r"(\.xlsx)$", r"_review\1", args.output)
        ws_review.parent.save(review_path)
        print(f"{review_row - 3} unidentified row(s) -> {review_path} "
              f"(fill in Artsnavn, import separately).")

    # Move handled recordings aside (in per-day folders, with their transcripts)
    # so the next run only sees new ones and days can be re-processed for free.
    # Done only after a successful save, so a crash leaves everything reprocessable.
    if not args.keep:
        for sub, items in (("processed", processed), ("skipped", skipped)):
            for path, tx in items:
                day = os.path.basename(path)[:10]  # YYYY-MM-DD prefix
                dest = os.path.join(args.recordings, sub, day)
                os.makedirs(dest, exist_ok=True)
                shutil.move(path, os.path.join(dest, os.path.basename(path)))
                if tx:
                    with open(os.path.join(dest, os.path.splitext(os.path.basename(path))[0] + ".txt"),
                              "w", encoding="utf-8") as f:
                        f.write(tx)
        print(f"Moved recordings into {args.recordings}/processed (and /skipped).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
