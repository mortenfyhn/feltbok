#!/usr/bin/env python3
"""Build import test sheets to pin down how Artsobservasjoner matches localities.

What the first round taught us (from the new site's validation):
  * Lokalitetsnavn must be the BARE name ("Ørndalen"), NOT the qualified
    "Ørndalen, Sistranda, Frøya, Tø" — the composite is never found.
  * A bare name that is non-unique fails with "Flere lokaliteter med samme navnet".
  * The new site (v3.0) has a separate "Hovedlokalitet" column to disambiguate;
    the old site (v2.20) does not.

So two open questions, one per site, tested here with bare names + the exact
registry coordinates (from GBIF/Artsdatabanken, the canonical site point):

  OLD site (v2.20, your preference): does bare name + EXACT coords link to the
  correct existing locality, or still mint a new one? (-> lokalitetstest-gammel.xlsx)

  NEW site (v3.0): does bare Lokalitetsnavn + Hovedlokalitet resolve uniquely?
  (-> lokalitetstest-ny.xlsx)

Paste each into its site's Importer observasjoner (do NOT publish) and read the
per-row validation. Each row's private comment says what it tests.

    python process/make_locality_test.py
"""

from openpyxl import load_workbook


def fugl_headers(wb):
    """Map column name -> 1-based column index from the Fugl header (row 2)."""
    ws = wb["Fugl"]
    return {str(c.value).strip(): c.column for c in ws[2] if c.value is not None}


def pick(cols, *names):
    """First of `names` present in the sheet (handles v2.20 vs v3.0 renames)."""
    for n in names:
        if n in cols:
            return n
    return names[0]


def write_row(ws, row, cols, values):
    for name, value in values.items():
        if value is None or value == "" or name not in cols:
            continue
        ws.cell(row=row, column=cols[name], value=value)


DATE = "31.05.2026"
SPECIES = "Kjøttmeis"
ACC = "100 m"

# Real Frøya localities with their exact registry coordinates (GBIF site point).
LOCS = [
    {
        "lok": "Ørndalen",
        "hoved": "Sistranda",
        "lat": 63.72865,
        "lon": 8.81729,
        "note": "tvetydig navn (flere i NO)",
    },
    {
        "lok": "Ellingsundet",
        "hoved": "Uttian",
        "lat": 63.76564,
        "lon": 8.82686,
        "note": "trolig unikt navn",
    },
    {
        "lok": "Flatvalsundet",
        "hoved": "Flatval",
        "lat": 63.68704,
        "lon": 8.75778,
        "note": "var «ikke funnet» bart",
    },
]


def writer(wb):
    cols = fugl_headers(wb)
    ws = wb["Fugl"]
    pub = pick(cols, "Merknad (synlig for alle)", "Kommentar (synlig for alle)")
    priv = pick(
        cols,
        "Privat merknad (kun synlig for deg selv)",
        "Privat kommentar (kun synlig for deg selv)",
    )

    def emit(row, name, label, *, hoved=None, coords=False):
        vals = {
            "Artsnavn": SPECIES,
            "Lokalitetsnavn": name,
            "Fra dato": DATE,
            "Til dato": DATE,
            "Fra klokkeslett": "08:00",
            "Til klokkeslett": "08:00",
            "Antall": 1,
            pub: "TEST – ikke publiser",
            priv: label,
        }
        if hoved:
            vals["Hovedlokalitet"] = hoved
        if coords:
            vals.update({"Nord": coords[0], "Øst": coords[1], "Nøyaktighet": ACC})
        write_row(ws, row, cols, vals)
        print(f"  L{row} {label}")

    return ws, emit


def build_old():
    """v2.20: bare name, with and without exact coords."""
    wb = load_workbook("docs/artsobs-template-v2.20.xlsx")
    _, emit = writer(wb)
    row = 3
    for L in LOCS:
        emit(row, L["lok"], f"bart, uten koord — {L['note']}")
        row += 1
        emit(
            row,
            L["lok"],
            "bart + eksakte koord — kobles til rett lokalitet eller ny?",
            coords=(L["lat"], L["lon"]),
        )
        row += 1
    out = "lokalitetstest-gammel.xlsx"
    wb.save(out)
    print(f"-> {out} (gammel/v2.20: bart navn ± eksakte koordinater)\n")


def build_new():
    """v3.0: bare Lokalitetsnavn + Hovedlokalitet, no coords."""
    wb = load_workbook("docs/artsobs-template-v3.0.xlsx")
    _, emit = writer(wb)
    row = 3
    for L in LOCS:
        emit(
            row, L["lok"], f"Lok + Hovedlok «{L['hoved']}» — entydig?", hoved=L["hoved"]
        )
        row += 1
    emit(row, "Ørndalen", "kontroll: Lok uten Hovedlok — forventer tvetydig")
    out = "lokalitetstest-ny.xlsx"
    wb.save(out)
    print(f"-> {out} (ny/v3.0: Lokalitetsnavn + Hovedlokalitet)")


def main():
    build_old()
    build_new()


if __name__ == "__main__":
    main()
